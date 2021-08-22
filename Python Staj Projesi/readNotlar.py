
import openpyxl, pprint, statistics #openpyxl modülü içeri aktarıldı.
from openpyxl.styles import Alignment  
print('Opening workbook...')
wb = openpyxl.load_workbook('Notlar.xlsx') #çalışma kitabını erişildi
sheet = wb.get_sheet_by_name('BIL2001') #çalışma kitabından BIL2001 Sayfası alındı

studentData = [] #aldığımız verileri aktarabilmek için bir liste oluşturduk


print('Reading rows...')

for row in range(1, sheet.max_row + 1):  # sheet.get_highest_row()
#sheet.max_row+1 -> satır sonuna kadar inebilmek için
    """
    blokta A - B - C sütunlarında bulunan değerleri satır sırasıyla 
    numara vize ve final değişkenlerine atayıp bir item oluşturuldu.
    """    
    numara = sheet['A' + str(row)].value
    vize   = sheet['B' + str(row)].value
    final  = sheet['C' + str(row)].value
    
    # -> Öğrenci Ortalamasını Hesapla (studentDAta)iteme ekle
    ortalama = (vize + final)/2
    
    
    print(str(numara) +'  '+ str(vize) +'  '+ str(final) +' '+ str(ortalama))

    item = [numara, vize, final, ortalama]
    studentData.insert(row, item) #(row, item) ifadesiyle for içinde bulunan satır sırasıyla aynı şekilde studentData listemize eklemiş bulunduk.





print('Writing results...') # sonuçlar yazıldığı kısma geçiyoruz

max_row = sheet.max_row # Notlar.xlsx
#satır sayısı max_row adlı değişkene atandı.

wb = openpyxl.Workbook() # bir çalışma kitabı açıldı
sheet = wb.get_sheet_by_name('Sheet') #sayfa ismine sheet verildi

# Hücre Birleştirmesi
sheet.merge_cells('A1:E1')
cell1 = sheet.cell(row=1, column=1)  
cell1.value = 'BAGIL DEGERLENDIRMEYE KATILMAYANLAR:'  
cell1.alignment = Alignment(horizontal='right', vertical='center')  

sheet.merge_cells('A2:E2')
cell2 = sheet.cell(row=2, column=1)  
cell2.value = 'SINIF ORTALAMASI:'
cell2.alignment = Alignment(horizontal='right', vertical='center')  

sheet.merge_cells('A3:E3')
cell3 = sheet.cell(row=3, column=1)  
cell3.value = 'STANDART SAPMA:'
cell3.alignment = Alignment(horizontal='right', vertical='center')  

sheet.merge_cells('A4:E4')
cell4 = sheet.cell(row=4, column=1)  
cell4.value = 'SINIF DUZEYI:'      
cell4.alignment = Alignment(horizontal='right', vertical='center')  
# ----------------------------------------------------------------------

# Başlıklar
sheet['A' + str(5)] = 'Numara'
sheet['B' + str(5)] = 'Vize'
sheet['C' + str(5)] = 'Final'
sheet['D' + str(5)] = 'Ortalama'
sheet['E' + str(5)] = 'T Notu'
sheet['F' + str(5)] = 'Harf Notu'
# -----------------------------------------------------------------------

for row in range(6, max_row + 6): #aldığımız max_row kadar fora sayma verdik.
# Böylelikle oluşturlan studentData sütun ve satır karıştırılmadan yeni bir excell sayfasına yazıldı.
    sheet['A' + str(row)] = studentData[row-6][0]
    sheet['B' + str(row)] = studentData[row-6][1]
    sheet['C' + str(row)] = studentData[row-6][2]
    sheet['D' + str(row)] = studentData[row-6][3]

# Sınıf Ortalaması ve Bağıl değerlendirmeye katılmayanlar
ortList = [] 
katilmayanlar= 0
for row in range(6 , sheet.max_row + 1):
    final = sheet['C' + str(row)].value
    ortalama = sheet['D' + str(row)].value
    if final == 0:
        katilmayanlar += 0
    elif final < 45:
        if ortalama > 15:
            ortList.insert(row, ortalama)
        katilmayanlar +=1
    else:
        ortList.insert(row, ortalama)

sinifOrtalamasi = sum(ortList)/len(ortList)

sheet['F' + str(2)] = round(sinifOrtalamasi , 2) 
sheet['F' + str(1)] = katilmayanlar
# -----------------------------------------------------------------------


# Sınıf Düzeyi Hesaplama
sinifDuzeyi=''
if sinifOrtalamasi <=42.5:
    sinifDuzeyi = 'Kötü'
elif 42.5 < sinifOrtalamasi <= 47.5:
    sinifDuzeyi = 'Zayıf'
elif 47.5 < sinifOrtalamasi <= 52.5:
    sinifDuzeyi = 'Orta'
elif 52.5 < sinifOrtalamasi <= 57.5:
    sinifDuzeyi = 'Ortanın Üstü'
elif 57.5 < sinifOrtalamasi <= 62.5:
    sinifDuzeyi = 'İyi'
elif  62.5 < sinifOrtalamasi <= 70:
    sinifDuzeyi = 'Çok İyi'
elif 70 < sinifOrtalamasi <= 80:
    sinifDuzeyi = 'Mükemmel'
else:
    sinifDuzeyi = 'Üstün Başarı'

sheet['F' + str(4)] = sinifDuzeyi
sheet['F' + str(4)].alignment = Alignment(horizontal='right', vertical = 'center')
# -----------------------------------------------------------------------

#Standart Sapma
varyans = 0
standartSapma = statistics.stdev(ortList)
sheet['F' + str(3)] = round(standartSapma , 2) 
# -----------------------------------------------------------------------

# T Notu Hesaplama  **********************************
tNot = 0.0
for row in range(6, sheet.max_row + 1):
    if sheet['D' + str(row)].value == 0:
        tNot = 0
    elif sheet['D' + str(row)].value == sinifOrtalamasi:
        tNot = 50 
    else: 
        zNot = (sheet['D' + str(row)].value - sinifOrtalamasi)/standartSapma
        tNot = (zNot*10)+50
        
    sheet['E' + str(row)] = round(tNot) 

# -----------------------------------------------------------------------

# Harf Notu
AA = 0 
BA = 0 
BB = 0 
CB = 0 
CC = 0 
DC = 0 
DD = 0 
FD = 0 
FF = 0

for row in range(6,sheet.max_row+1):
    tNot = float(sheet['E' + str(row)].value)
    if sinifDuzeyi == 'Kötü':
        if tNot < 36:
            harfNot = 'FF'
        elif 36 <= tNot <= 40.99:
            harfNot = 'FD'
        elif 41 <= tNot <= 45.99:
            harfNot = 'DD'
        elif 46 <= tNot <= 50.99:
            harfNot = 'DC'
        elif 51 <= tNot <= 55.99:
            harfNot = 'CC'
        elif 56 <= tNot <= 60.99:
            harfNot = 'CB'
        elif 61 <= tNot <= 65.99:
            harfNot = 'BB'
        elif 66 <= tNot <= 70.99:
            harfNot = 'BA'
        else:
            harfNot = 'AA'
            
    if sinifDuzeyi == 'Zayıf':
        if tNot < 34:
            harfNot = 'FF'
        elif 34 <= tNot <= 38.99:
            harfNot = 'FD'
        elif 39 <= tNot <= 43.99:
            harfNot = 'DD'
        elif 44 <= tNot <= 48.99:
            harfNot = 'DC'
        elif 49 <= tNot <= 53.99:
            harfNot = 'CC'
        elif 54 <= tNot <= 58.99:
            harfNot = 'CB'
        elif 59 <= tNot <= 63.99:
            harfNot = 'BB'
        elif 64 <= tNot <= 68.99:
            harfNot = 'BA'
        else:
            harfNot = 'AA'
            
    if sinifDuzeyi == 'Orta':
        if tNot < 32:
            harfNot = 'FF'
        elif 32 <= tNot <= 36.99:
            harfNot = 'FD'
        elif 37 <= tNot <= 41.99:
            harfNot = 'DD'
        elif 42 <= tNot <= 46.99:
            harfNot = 'DC'
        elif 47 <= tNot <= 51.99:
            harfNot = 'CC'
        elif 52 <= tNot <= 56.99:
            harfNot = 'CB'
        elif 57 <= tNot <= 61.99:
            harfNot = 'BB'
        elif 62 <= tNot <= 66.99:
            harfNot = 'BA'
        else:
            harfNot = 'AA'
            
    if sinifDuzeyi == 'Ortanın Üstü':
        if tNot < 30:
            harfNot = 'FF'
        elif 30 <= tNot <= 34.99:
            harfNot = 'FD'
        elif 35 <= tNot <= 39.99:
            harfNot = 'DD'
        elif 40 <= tNot <= 44.99:
            harfNot = 'DC'
        elif 45 <= tNot <= 49.99:
            harfNot = 'CC'
        elif 50 <= tNot <= 54.99:
            harfNot = 'CB'
        elif 55 <= tNot <= 59.99:
            harfNot = 'BB'
        elif 60 <= tNot <= 64.99:
            harfNot = 'BA'
        else:
            harfNot = 'AA'
                
    if sinifDuzeyi == 'İyi':
        if tNot < 28:
            harfNot = 'FF'
        elif 28 <= tNot <= 32.99:
            harfNot = 'FD'
        elif 33 <= tNot <= 37.99:
            harfNot = 'DD'
        elif 38 <= tNot <= 42.99:
            harfNot = 'DC'
        elif 43 <= tNot <= 47.99:
            harfNot = 'CC'
        elif 48 <= tNot <= 52.99:
            harfNot = 'CB'
        elif 53 <= tNot <= 57.99:
            harfNot = 'BB'
        elif 58 <= tNot <= 62.99:
            harfNot = 'BA'
        else:
            harfNot = 'AA'
                
    if sinifDuzeyi == 'Çok İyi':
        if tNot < 26:
            harfNot = 'FF'
        elif 26 <= tNot <= 30.99:
            harfNot = 'FD'
        elif 31 <= tNot <= 35.99:
            harfNot = 'DD'
        elif 36 <= tNot <= 40.99:
            harfNot = 'DC'
        elif 41 <= tNot <= 45.99:
            harfNot = 'CC'
        elif 46 <= tNot <= 50.99:
            harfNot = 'CB'
        elif 51 <= tNot <= 55.99:
            harfNot = 'BB'
        elif 56 <= tNot <= 60.99:
            harfNot = 'BA'
        else:
            harfNot = 'AA'
            
    if sinifDuzeyi == 'Mükemmel':
        if tNot < 24:
            harfNot = 'FF'
        elif 24 <= tNot <= 28.99:
            harfNot = 'FD'
        elif 29 <= tNot <= 33.99:
            harfNot = 'DD'
        elif 34 <= tNot <= 38.99:
            harfNot = 'DC'
        elif 39 <= tNot <= 43.99:
            harfNot = 'CC'
        elif 44 <= tNot <= 48.99:
            harfNot = 'CB'
        elif 49 <= tNot <= 53.99:
            harfNot = 'BB'
        elif 54 <= tNot <= 58.99:
            harfNot = 'BA'
        else:
            harfNot = 'AA'
    
    sheet['F' + str(row)] = harfNot
    sheet['F' + str(row)].alignment = Alignment(horizontal='right', vertical='center')
# -----------------------------------------------------------------------    
# Harf notu saydırma 
    if harfNot == 'AA':
        AA += 1
    if harfNot == 'BA':
        BA += 1
    if harfNot == 'BB':
        BB += 1
    if harfNot == 'CB':
        CB += 1
    if harfNot == 'CC':
        CC += 1
    if harfNot == 'DC':
        DC += 1
    if harfNot == 'DD':
        DD += 1 
    if harfNot == 'FD':
        FD += 1
    if harfNot == 'FF':
        FF += 1
# -----------------------------------------------------------------------   
# Hücre Birleştirme ve Harf 
sheet.merge_cells('I1:M1')
cell = sheet.cell(row=1, column=9)  
cell.value = 'HARF NOTLARI DAĞILIMI:'      
cell.alignment = Alignment(horizontal='center', vertical='center')  

harfNotlari = ['AA', 'BA', 'BB', 'CB', 'CC', 'DC', 'DD', 'FD', 'FF']

sayac = 2
for harf in harfNotlari:
    sheet['J' + str(sayac)] = harf
    sayac += 1

sheet['K' + str(2)].value = AA
sheet['K' + str(3)].value = BA
sheet['K' + str(4)].value = BB
sheet['K' + str(5)].value = CB
sheet['K' + str(6)].value = CC
sheet['K' + str(7)].value = DC
sheet['K' + str(8)].value = DD
sheet['K' + str(9)].value = FD
sheet['K' + str(10)].value = FF
    


# -----------------------------------------------------------------

 




       
wb.save('Notlar_.xlsx') #Notlar_.xlsx adı ile kaydedildi.
print('Done.') #işlemler bitti.


